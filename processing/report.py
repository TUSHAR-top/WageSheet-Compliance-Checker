"""
Build the downloadable verification report as an .xlsx workbook.

Everything lives on one worksheet (summary block, then the per-employee
table) rather than split across tabs. Some viewers (Explorer's preview pane,
Outlook's attachment preview, etc.) only render the active sheet and hide
tab-switching entirely, which would silently drop a second sheet from view.
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .verify import (STATUS_AMOUNT_MISMATCH, STATUS_MATCHED,
                      STATUS_MISSING, STATUS_NAME_MISMATCH)

STATUS_FILLS = {
    STATUS_MATCHED: PatternFill("solid", fgColor="C6EFCE"),
    STATUS_NAME_MISMATCH: PatternFill("solid", fgColor="FFEB9C"),
    STATUS_AMOUNT_MISMATCH: PatternFill("solid", fgColor="FFD9B3"),
    STATUS_MISSING: PatternFill("solid", fgColor="FFC7CE"),
}
HEADER_FILL = PatternFill("solid", fgColor="2F5597")
HEADER_FONT = Font(color="FFFFFF", bold=True)
STATUS_ORDER = [STATUS_MATCHED, STATUS_NAME_MISMATCH, STATUS_AMOUNT_MISMATCH, STATUS_MISSING]

DETAIL_HEADERS = [
    "Row", "Name (Wagesheet)", "UAN", "IP Number", "Agency", "State",
    "PF Status", "PF Wagesheet Amt", "PF ECR Amt", "PF Diff", "PF ECR Name", "PF Reason",
    "ESIC Status", "ESIC Wagesheet Amt", "ESIC ECR Amt", "ESIC Diff", "ESIC ECR Name", "ESIC Reason",
]
DETAIL_WIDTHS = [6, 22, 15, 15, 22, 12, 15, 14, 12, 9, 20, 34, 15, 14, 12, 9, 20, 34]


def build_report(results, summary, warnings=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Verification Report"

    next_row = _write_summary_block(ws, summary, warnings or [])
    _write_details_table(ws, results, start_row=next_row)

    for col_idx, w in enumerate(DETAIL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    return wb


def _write_summary_block(ws, summary, warnings):
    ws["A1"] = "Wagesheet vs. ECR Compliance Verification"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = "Total employees in wagesheet"
    ws["B2"] = summary["total_employees"]

    row = 4
    for label, key in (("PF (vs. PF ECR)", "pf"), ("ESIC (vs. ESIC ECR)", "esic")):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        row += 1
        for col, status in enumerate(STATUS_ORDER, start=1):
            c = ws.cell(row=row, column=col, value=status)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
        row += 1
        counts = summary[key]
        for col, status in enumerate(STATUS_ORDER, start=1):
            c = ws.cell(row=row, column=col, value=counts.get(status, 0))
            c.fill = STATUS_FILLS.get(status)
            c.alignment = Alignment(horizontal="center")
        row += 2

    if warnings:
        ws.cell(row=row, column=1, value="Parsing warnings").font = Font(bold=True)
        row += 1
        for w in warnings:
            ws.cell(row=row, column=1, value=w)
            row += 1

    return row + 1


def _write_details_table(ws, results, start_row):
    header_row = start_row
    ws.cell(row=header_row, column=1, value="Per-employee results").font = Font(bold=True, size=12)

    header_row += 1
    for col_idx, title in enumerate(DETAIL_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    first_data_row = header_row + 1
    for offset, r in enumerate(results):
        r_idx = first_data_row + offset
        pf, esic = r["pf"], r["esic"]
        values = [
            r["row_no"], r["name"], r["uan"], r["ip_number"], r["agency"], r["state"],
            pf["status"], pf["wagesheet_amount"], pf["ecr_amount"], pf["diff"], pf["ecr_name"], pf["reason"],
            esic["status"], esic["wagesheet_amount"], esic["ecr_amount"], esic["diff"], esic["ecr_name"], esic["reason"],
        ]
        for c_idx, value in enumerate(values, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        ws.cell(row=r_idx, column=7).fill = STATUS_FILLS.get(pf["status"])
        ws.cell(row=r_idx, column=13).fill = STATUS_FILLS.get(esic["status"])

    last_row = first_data_row + len(results) - 1
    if last_row >= header_row:
        ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(DETAIL_HEADERS))}{last_row}"
    ws.freeze_panes = f"A{first_data_row}"
